import streamlit as st
import pandas as pd
import io
import time
from datetime import datetime
from utils.helpers import sanitize_sheet_name, safe_dataframe
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def render_tab(container, supabase, username, role, loc_list, t):
    with container:
        st.header(t["inventory_header"])

        # --- ENTRY LOGIC ---
        try:
            query = supabase.table("Master_Inventory").select("*").execute()
            master_inventory = pd.DataFrame(query.data) if query.data else pd.DataFrame()
        except Exception:
            master_inventory = pd.DataFrame()

        if not master_inventory.empty:
            # Location filter
            if role in ["Admin", "Manager"]:
                sel_loc = st.selectbox(
                    t["location"],
                    ["All Locations"] + sorted(master_inventory['Location'].unique()),
                    key="inventory_location_select"
                )
                inv_df = master_inventory if sel_loc == "All Locations" else master_inventory[master_inventory['Location'] == sel_loc]
            else:
                inv_df = master_inventory[master_inventory['Location'] == loc_list[0]]
                st.write(f"📍 {t['location']}: {', '.join(loc_list)}")
                sel_loc = loc_list[0]

            # Category selection
            sel_cat = st.selectbox(
                "Select Category",
                sorted(inv_df['Category'].unique()),
                key="inventory_category_select"
            )
            cat_df = inv_df[inv_df['Category'] == sel_cat].copy()

            # Editable physical counts
            cat_df['Total_Physical'] = 0
            edited_df = st.data_editor(
                cat_df[['Full Name','SKU','Stock','Total_Physical']],
                num_rows="dynamic",
                width='stretch',
                key="inventory_data_editor"
            )

            # Save audit entries
            if st.button("✅ Save Audit", key="inventory_save_audit"):
                for _, row in edited_df.iterrows():
                    audit_entry = {
                        "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "Name": row['Full Name'],
                        "Category": sel_cat,
                        "System_Stock": row['Stock'],
                        "Total_Physical": row['Total_Physical'],
                        "Discrepancy": row['Total_Physical'] - row['Stock'],
                        "Counter_Name": username,
                        "location": sel_loc
                    }
                    supabase.table("Inventory").insert(audit_entry).execute()
                st.success("Audit saved successfully!")
                time.sleep(1); st.rerun()
        else:
            st.info("No data in Master_Inventory.")

        st.divider(); st.subheader("📜 Audit History by Category")

        # --- HISTORY + EXCEL EXPORT ---
        try:
            aud_log_res = supabase.table("Inventory").select("*").order("Date", desc=True).execute()
            if aud_log_res.data:
                df_log = pd.DataFrame(aud_log_res.data)

                # Filter by location
                if role == "Staff":
                    df_log = df_log[df_log['location'].isin(loc_list)]
                elif role in ["Admin","Manager"] and sel_loc != "All Locations":
                    df_log = df_log[df_log['location'] == sel_loc]

                output = io.BytesIO()
                summary_rows = []
                
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    for cat in sorted(df_log['Category'].unique()):
                        cat_df = df_log[df_log['Category'] == cat].copy()
                
                        export_df = cat_df.rename(columns={
                            "Name": "Nom",
                            "Category": "Category",
                            "Counter_Name": "Employé",
                            "Total_Physical": "Total Physique",
                            "System_Stock": "Système",
                            "Discrepancy": "Différence",
                            "location": "Local"
                        })
                
                        safe_name = sanitize_sheet_name(cat)
                
                        # Leave space for header (we start writing at row 4)
                        export_df.to_excel(writer, sheet_name=safe_name, index=False, startrow=3)
                
                        summary_rows.append({
                            "Catégorie": cat,
                            "Sheet Name": safe_name,
                            "Total Compté": export_df["Total Physique"].sum(),
                            "Total System": export_df["Système"].sum(),
                            "Total Différence": export_df["Différence"].sum()
                        })
                
                summary_df = pd.DataFrame(summary_rows)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # ---------------- POST STYLING ----------------
                output.seek(0)
                wb = load_workbook(output)
                
                from openpyxl.styles import Alignment, Font, Border, Side
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for sheet_name in wb.sheetnames:
                    if sheet_name == "Summary":
                        continue
                
                    ws = wb[sheet_name]
                    max_col = ws.max_column
                    max_row = ws.max_row
                
                    # === 1. DATE HEADER ===
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                    cell = ws.cell(row=1, column=1)
                    cell.value = datetime.now().strftime("%d-%m-%Y")
                    cell.font = Font(size=16, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                    ws.row_dimensions[1].height = 25
                
                    # === 2. COLUMN HEADERS STYLE ===
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=4, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                
                    # === 3. DATA CELLS STYLE ===
                    for row in range(5, max_row + 1):
                        for col in range(1, max_col + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = thin_border
                
                    # === 4. LEFT VERTICAL CATEGORY LABEL ===
                    start_row = 4
                    end_row = max_row
                
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    vcell = ws.cell(row=start_row, column=1)
                
                    vcell.value = sheet_name.upper()
                    vcell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
                    vcell.font = Font(bold=True)
                
                    # Shift table right (since col 1 is used)
                    for row in range(4, max_row + 1):
                        for col in reversed(range(2, max_col + 2)):
                            ws.cell(row=row, column=col).value = ws.cell(row=row, column=col - 1).value
                
                    # Clear duplicated col 1 (except vertical label)
                    for row in range(5, max_row + 1):
                        ws.cell(row=row, column=1).value = None
                
                    # === 5. FOOTER ===
                    footer_row = max_row + 2
                
                    ws.cell(row=footer_row, column=1).value = "EXPLICATIONS:"
                    ws.cell(row=footer_row + 2, column=1).value = "DRESSUP HAITI - INVENTAIRE PÉTION-VILLE"
                
                # Save final file
                final_output = io.BytesIO()
                wb.save(final_output)
                final_output.seek(0)
                
                st.download_button(
                    "⬇️ Download Full Audit History (Styled Excel Report)",
                    data=final_output.getvalue(),
                    file_name="audit_history_by_category.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="inventory_download_button"
                )

                st.divider(); st.subheader("📊 Summary Preview")
                st.dataframe(summary_df, width='stretch', hide_index=True, key="inventory_summary_preview")
            else:
                st.info("No audit records found yet.")
        except Exception as e:
            st.error(f"Error fetching history: {e}")
